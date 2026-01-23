import csv
import io
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExportService:
    """Service for exporting invoice data to various formats."""
    
    @staticmethod
    def _get_val(obj, key, default=None):
        """Helper to get value from either object or dictionary."""
        if hasattr(obj, key):
            return getattr(obj, key)
        if isinstance(obj, dict):
            return obj.get(key, default)
        return default

    @staticmethod
    def invoices_to_dict_list(invoices: List[Any]) -> List[Dict[str, Any]]:
        """Convert invoice objects or dicts to list of dictionaries."""
        result = []
        for inv in invoices:
            created_at = ExportService._get_val(inv, "created_at")
            if isinstance(created_at, str):
                try:
                    created_at = datetime.fromisoformat(created_at)
                except ValueError:
                    pass

            row = {
                "ID": ExportService._get_val(inv, "id"),
                "Invoice No": ExportService._get_val(inv, "invoice_number") or "",
                "Date": ExportService._get_val(inv, "invoice_date") or "",
                "Supplier": ExportService._get_val(inv, "supplier_name") or "",
                "Amount": ExportService._get_val(inv, "total_amount") or 0,
                "Currency": ExportService._get_val(inv, "currency") or "TRY",
                "Tax Amount": ExportService._get_val(inv, "tax_amount") or 0,
                "Tax Rate (%)": ExportService._get_val(inv, "tax_rate") or 0,
                "Status": ExportService._get_val(inv, "status"),
                "Processing Time (ms)": ExportService._get_val(inv, "processing_time_ms") or 0,
                "Created At": created_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(created_at, datetime) else str(created_at or ""),
                "File Name": ExportService._get_val(inv, "original_filename") or "",
                "File Type": ExportService._get_val(inv, "file_type") or "",
            }
            result.append(row)
        return result
    
    @staticmethod
    def export_to_csv(invoices: List[Any]) -> str:
        """Export invoices to CSV format."""
        data = ExportService.invoices_to_dict_list(invoices)
        
        if not data:
            return ""
        
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
        
        return output.getvalue()
    
    @staticmethod
    def export_to_excel(invoices: List[Any], include_items: bool = True) -> bytes:
        """Export invoices to Excel format with styling."""
        wb = Workbook()
        
        # ===== Invoices Sheet =====
        ws = wb.active
        ws.title = "Invoices"
        
        data = ExportService.invoices_to_dict_list(invoices)
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
            
        if data:
            # Write headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, (key, value) in enumerate(row_data.items(), 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    
                    # Format numbers
                    if isinstance(value, (int, float)) and key in ["Amount", "Tax Amount"]:
                        cell.number_format = '#,##0.00'
            
            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                max_length = max(
                    len(str(ws.cell(row=row, column=col).value or "")) 
                    for row in range(1, len(data) + 2)
                )
                ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)
        
        # ===== Line Items Sheet =====
        if include_items:
            ws_items = wb.create_sheet("Items")
            
            item_headers = [
                "Invoice ID", "Invoice No", "Product Name", "Description",
                "Quantity", "Unit Price", "Total Price", "Validation"
            ]
            
            # Write headers
            for col, header in enumerate(item_headers, 1):
                cell = ws_items.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Write items
            row_idx = 2
            for inv in invoices:
                items = ExportService._get_val(inv, "items", [])
                inv_id = ExportService._get_val(inv, "id")
                inv_no = ExportService._get_val(inv, "invoice_number") or ""
                
                for item in items:
                    ws_items.cell(row=row_idx, column=1, value=inv_id)
                    ws_items.cell(row=row_idx, column=2, value=inv_no)
                    ws_items.cell(row=row_idx, column=3, value=ExportService._get_val(item, "product_name", ""))
                    ws_items.cell(row=row_idx, column=4, value=ExportService._get_val(item, "description", ""))
                    ws_items.cell(row=row_idx, column=5, value=ExportService._get_val(item, "quantity", 0))
                    ws_items.cell(row=row_idx, column=6, value=ExportService._get_val(item, "unit_price", 0))
                    ws_items.cell(row=row_idx, column=7, value=ExportService._get_val(item, "total_price", 0))
                    
                    is_valid = ExportService._get_val(item, "is_arithmetic_valid")
                    ws_items.cell(row=row_idx, column=8, value="✓" if is_valid else "✗")
                    
                    for col in range(1, 9):
                        ws_items.cell(row=row_idx, column=col).border = thin_border
                    
                    row_idx += 1
            
            # Auto-adjust column widths
            for col in range(1, len(item_headers) + 1):
                ws_items.column_dimensions[get_column_letter(col)].width = 15
        
        # ===== Summary Sheet =====
        ws_summary = wb.create_sheet("Summary")
        
        summary_data = [
            ("Total Invoice Count", len(invoices)),
            ("Successful Jobs", sum(1 for i in invoices if ExportService._get_val(i, "status") == "completed")),
            ("Failed Jobs", sum(1 for i in invoices if ExportService._get_val(i, "status") == "failed")),
            ("Total Amount", sum(ExportService._get_val(i, "total_amount") or 0 for i in invoices)),
            ("Total Tax", sum(ExportService._get_val(i, "tax_amount") or 0 for i in invoices)),
            ("Average Processing Time (ms)",
             sum(ExportService._get_val(i, "processing_time_ms") or 0 for i in invoices) / len(invoices) if invoices else 0),
            ("Report Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            cell = ws_summary.cell(row=row_idx, column=2, value=value)
            if isinstance(value, float):
                cell.number_format = '#,##0.00'
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
